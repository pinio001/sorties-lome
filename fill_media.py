"""
fill_media.py
-------------
Scanne les dossiers ./places_files ou ./events_files
et remplit la colonne 'media' (places) ou 'image' (events)
dans le fichier Excel avec les URLs GitHub raw correspondantes.

Usage :
    python fill_media.py places   -> remplit template_places.xlsx colonne 'media'
    python fill_media.py events   -> remplit template_events.xlsx  colonne 'image'

Configuration :
    Modifie GITHUB_USER, GITHUB_REPO et GITHUB_BRANCH selon ton repo.
"""

import sys
import os
import re
import pandas as pd
from openpyxl import load_workbook

# ─── CONFIG — À adapter à ton repo ───────────────────────────────────────────
GITHUB_USER   = "pinio001"
GITHUB_REPO   = "sorties-lome"
GITHUB_BRANCH = "main"
IMAGE_EXTS    = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
# ─────────────────────────────────────────────────────────────────────────────

def slugify(name: str) -> str:
    name = name.strip()
    name = re.sub(r"[^\w\s\-]", "", name, flags=re.UNICODE)
    name = re.sub(r"[\s]+", "_", name)
    return name

def get_images(folder_path: str, max_count: int) -> list[str]:
    """Retourne les fichiers image triés (1, 2, 3...) dans un dossier."""
    if not os.path.isdir(folder_path):
        return []
    files = []
    for f in os.listdir(folder_path):
        ext = os.path.splitext(f)[1].lower()
        if ext in IMAGE_EXTS:
            files.append(f)
    # Trier par nom numérique
    files.sort(key=lambda x: int(re.sub(r"\D", "", os.path.splitext(x)[0]) or "0"))
    return files[:max_count]

def build_url(subfolder: str, folder_name: str, filename: str) -> str:
    return (
        f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/"
        f"{GITHUB_BRANCH}/{subfolder}/{folder_name}/{filename}"
    )

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ("places", "events"):
        print("Usage : python fill_media.py [places|events]")
        sys.exit(1)

    mode = sys.argv[1]

    if mode == "places":
        excel_file  = "template_places.xlsx"
        files_dir   = "places_files"
        name_col    = "name"
        media_col   = "media"
        max_images  = 4
    else:
        excel_file  = "template_events.xlsx"
        files_dir   = "events_files"
        name_col    = "title"
        media_col   = "image"
        max_images  = 1

    if not os.path.exists(excel_file):
        print(f"❌ Fichier introuvable : {excel_file}")
        sys.exit(1)

    if not os.path.exists(files_dir):
        print(f"❌ Dossier introuvable : ./{files_dir}/")
        print(f"   Lance d'abord : python create_folders.py {mode}")
        sys.exit(1)

    # ── Lire le fichier Excel avec openpyxl pour modifier en place ────────────
    wb = load_workbook(excel_file)
    ws = wb.active

    # Trouver les index de colonnes (ligne 3 = row 3 dans openpyxl)
    header_row = 3
    col_index = {}
    for cell in ws[header_row]:
        if cell.value:
            clean = str(cell.value).replace(" *", "").replace("\n(auto)", "").strip()
            col_index[clean] = cell.column

    if name_col not in col_index:
        print(f"❌ Colonne '{name_col}' non trouvée dans les en-têtes.")
        print(f"   En-têtes détectés : {list(col_index.keys())}")
        sys.exit(1)

    if media_col not in col_index:
        print(f"❌ Colonne '{media_col}' non trouvée dans les en-têtes.")
        sys.exit(1)

    name_col_idx  = col_index[name_col]
    media_col_idx = col_index[media_col]

    # ── Parcourir les lignes de données (à partir de row 5, row 4 = exemple) ──
    updated  = 0
    no_image = 0
    skipped  = 0

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        name_cell  = row[name_col_idx - 1]
        media_cell = row[media_col_idx - 1]

        raw_name = name_cell.value
        if not raw_name or str(raw_name).strip() == "":
            continue

        name       = str(raw_name).strip()
        folder_name = slugify(name)
        folder_path = os.path.join(files_dir, folder_name)

        images = get_images(folder_path, max_images)

        if not images:
            print(f"  ⚠️  Pas d'image : {folder_path}")
            no_image += 1
            continue

        if max_images == 1:
            # Events : 1 seule image
            url = build_url(files_dir, folder_name, images[0])
            media_cell.value = url
        else:
            # Places : plusieurs images séparées par |
            urls = [build_url(files_dir, folder_name, img) for img in images]
            media_cell.value = " | ".join(urls)

        print(f"  ✅ {name} → {len(images)} image(s)")
        updated += 1

    wb.save(excel_file)

    print(f"\n{'─'*50}")
    print(f"✅ {updated} ligne(s) mises à jour dans {excel_file}")
    print(f"⚠️  {no_image} lieu/event sans image (colonne laissée vide)")
    print(f"\n👉 Prochaines étapes :")
    print(f"   1. git add ./{files_dir}")
    print(f"   2. git commit -m \"add images {mode}\"")
    print(f"   3. git push")
    print(f"   4. Importer {excel_file} sur /admin/import")

if __name__ == "__main__":
    main()
